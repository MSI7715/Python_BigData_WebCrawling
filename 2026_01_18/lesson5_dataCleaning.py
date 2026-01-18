import requests
import openpyxl
import os
from bs4 import BeautifulSoup

# Excel活頁簿寫入
current_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join( current_dir, "job1111.xlsx" )

wb = openpyxl.Workbook()
ws = wb.active
ws[ "A1" ] = '職缺名稱'
ws[ "B1" ] = '公司名稱'
ws[ "C1" ] = '職缺連結'
ws[ "D1" ] = '工作地區'
ws[ "E1" ] = '縣市'
ws[ "F1" ] = '鄉鎮市區'
ws[ "G1" ] = '薪資待遇'
ws[ "H1" ] = '給薪方式'
ws[ "I1" ] = '薪資上限(月薪)'
ws[ "J1" ] = '薪資下限(月薪)'
ws[ "K1" ] = '薪資平均(月薪)'
ws[ "L1" ] = '經度'
ws[ "M1" ] = '緯度'
wb.save( file_path )

page = 1
res = requests.get(f'https://www.1111.com.tw/search/job?page={page}&col=ab&sort=desc&ks=%E5%A4%A7%E6%95%B8%E6%93%9A', verify=False)
soup = BeautifulSoup(res.text)
while soup( "div", class_ = "job-card" ) != [] and page < 15:
    for jobs in soup.find_all( "div", class_ = "job-card" ): # 職缺清單
        A = jobs( "h2" )[0].text
        B = jobs( "h2" )[1].text
        C = "https://www.1111.com.tw" + jobs( "a" )[0]["href"]
        D = jobs( "a" )[2].text
        E = jobs( "a" )[2].text[:3]
        F = jobs( "a" )[2].text[3:]
        G = jobs( "h4" )[0].text
        H = jobs( "h4" )[0].text[:2]
        pay_type = jobs( "h4" )[0].text[:2]
        salary = ""
        for st in jobs( "h4" )[0].text:
            if st.isdigit() or st == "~":
                salary += st
        print( "薪資範圍：", salary )
        # 如果~存在薪資範圍
        if "~" in salary:
            salary_min = int( salary[ :salary.find("~")] )
            salary_max = int( salary[ salary.find("~") + 1:] )
            salary_avg = ( salary_min + salary_max ) // 2
        else:
            salary_min = salary_max = salary_avg = int( salary )
        if pay_type == "面議":
            salary_min = salary_max = salary_avg = 40000
        elif pay_type == "月薪":
            pass
        elif pay_type ==  "年薪":
            salary_min /= 12
            salary_max /= 12
            salary_avg /= 12
        elif pay_type == "時薪":
            salary_min *= 21 * 8
            salary_max *= 21 * 8
            salary_avg *= 21 * 8
        else:
            print( pay_type )
        print( "薪資上限(月薪)：", salary_max )
        print( "薪資下限(月薪)：", salary_min )
        print( "平均薪資(月薪)：", salary_avg )
        ws.append( [ A, B, C, D, E, F, G, H, salary_max, salary_min, salary_avg ] )
    print("page", page)
    page += 1
    print()
    res = requests.get(f'https://www.1111.com.tw/search/job?page={page}&col=ab&sort=desc&ks=%E5%A4%A7%E6%95%B8%E6%93%9A', verify=False)
    soup = BeautifulSoup(res.text)
wb.save( file_path )